# -*- coding:utf8 -*-
import datetime

import pandas as pd
from openpyxl.styles import Alignment, Border, Side, numbers
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl import Workbook, load_workbook
from os import remove, path
import json


class XlsxSaver:
    """
    一个将DataFrame转换成格式化excel的工具
    """

    # 颜色
    Color_green = ['c6efce', '006100']  # 绿
    Color_red = ['ffc7ce', '9c0006']  # 红
    Color_yellow = ['ffeb9c', '9c6500']  # 黄
    Color_white = ['ffffff', '000000']  # 白底黑字
    Color_blue = ['ffffff', 'b8d9f3']  # 蓝底黑字

    def __init__(self, columns, df_in, filename='a.xlsx', sheet_name='Sheet1'):
        """
        df_in : 从一个DataFrame对象获取表格内容
        filename : 文件名
        sheet_name : 表名
        """
        self.columns = columns
        self.filename = filename  # 保存的xlsx文件的名字
        self.user_def = []  # 储存由用户自定义的列的列名，这些列不再参与自动计算列宽
        if path.exists(filename):
            # 如果文件存在，就直接打开，添加Sheet
            self.wb = load_workbook(filename)
            self.sheet = self.wb.create_sheet(sheet_name)
        else:
            # 如果文件不存在，就创建表格
            self.wb = Workbook()
            self.sheet = self.wb.active
            self.sheet.title = sheet_name
        # 将df的内容复制给sheet
        self.df = df_in.copy()
        # 插入序列
        self.df.insert(0, '序号', range(1, 1 + len(self.df)))
        self.columns.insert(0, '序号')
        # 设置表头
        self.sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(self.df.columns))
        self.sheet.cell(1, 1).value = sheet_name
        self.sheet.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')  # 居中

        # self.sheet.append(list(self.df.columns))
        self.sheet.append(list(self.columns))
        self.sheet.row_dimensions[1].height = 30  # 第一行高度
        self.sheet.row_dimensions[2].height = 20  # 第二行高度
        # 写入数据
        for row in range(0, len(list(self.df.index))):
            for col in range(0, len(list(self.df.columns))):
                self.sheet.cell(row + 3, col + 1).value = self.df.iloc[row, col]  # 注意：sheet行列从1开始计数
                self.sheet.cell(row + 3, col + 1).alignment = Alignment(horizontal='center', vertical='center')  # 居中

    def remove_file(self):
        remove(self.filename)

    def set_sheet_name(self, sheet_name):
        self.sheet.title = sheet_name

    def set_filename(self, filename):
        self.filename = filename

    def get_maxlength(self, series_in, col):
        """
        获取一个类型为object的Series中的最大占位长度，用于确定导出的xlsx文件的列宽
        col : 表头，也参与比较，解决有时候表头过长的问题
        """
        series = series_in.fillna('-')  # 填充空值，防止出现nan
        str_list = list(series)
        len_list = []
        for elem in str_list + [col]:
            elem_split = list(elem)
            length = 0
            for c in elem_split:
                if ord(c) <= 256:
                    length += 1
                else:
                    length += 2
            len_list.append(length)
        return max(len_list)

    def __auto_width(self):
        cols_list = list(self.df.columns)  # 获取列名
        for i in range(0, len(cols_list)):
            col = cols_list[i]
            if col in self.user_def:
                continue
            self.sheet.cell(1, i + 1).font = Font(bold=True, size=20)  # 加粗表头
            self.sheet.cell(2, i + 1).font = Font(bold=True, color=self.Color_yellow[1], size=12)  # 加粗列名,字体颜色
            self.sheet.cell(2, i + 1).alignment = Alignment(horizontal='center', vertical='center')  # 居中列名
            self.sheet.cell(2, i + 1).fill = PatternFill(fill_type="solid", start_color=self.Color_yellow[0],
                                                         end_color=self.Color_yellow[0])
            # 设置边框
            border = Border(
                left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000')
            )
            self.sheet.cell(2, i + 1).border = border

            letter = chr(i + 65)  # 由ASCII值获得对应的列字母
            max_len = self.get_maxlength(self.df[col].astype(str), col)
            if max_len <= 12:
                self.sheet.column_dimensions[letter].width = 12
            elif max_len <= 50:
                self.sheet.column_dimensions[letter].width = max_len + 2
            else:
                self.sheet.column_dimensions[letter].width = 50
                for cell in self.sheet[letter]:
                    cell.alignment = Alignment(wrap_text=True)

    def set_width(self, col_name, width):
        # 提供调整列宽的接口
        try:
            index = list(self.columns).index(col_name)
        except:
            return
        letter = chr(index + 65)
        self.sheet.column_dimensions[letter].width = width
        self.user_def.append(col_name)

    def set_color(self, col_name, color, rule):
        # 提供设置颜色的接口，rule:规则函数
        try:
            index = list(self.columns).index(col_name)
        except:
            return
        letter = chr(index + 65)
        for cell in self.sheet[letter]:
            if cell.row > 2 and rule(float(cell.value)):
                cell.fill = PatternFill(fill_type="solid", start_color=color[0], end_color=color[0])
                cell.font = Font(color=color[1])

    def set_percent(self, col_name):
        # 提供设置百分比的接口
        try:
            index = list(self.columns).index(col_name)
        except:
            return
        letter = chr(index + 65)
        for cell in self.sheet[letter]:
            if cell.row > 2:
                cell.number_format = numbers.FORMAT_PERCENTAGE

    def set_center_alignment(self, col_name):
        try:
            index = list(self.columns).index(col_name)
        except:
            return
        letter = chr(index + 65)
        for cell in self.sheet[letter]:
            cell.alignment = Alignment(wrap_text=True, horizontal='center')

    def save(self):
        # 自动调整列宽，并保存
        self.__auto_width()
        self.wb.save(self.filename)

    def set_merge(self, col_name):
        self.user_def.append(col_name)  # 设置为自定义列
        # 设置一列合并单元格
        try:
            index = list(self.columns).index(col_name)
        except:
            return
        letter = chr(index + 65)
        i = 1
        while True:
            if i >= self.sheet.max_row:
                # 结束条件：单元格到底
                break
            cell = self.sheet[letter + str(i)]
            j = i + 1  # 第一步指向下一个单元格
            while True:
                # 这个过程对j进行试探，最终j指向的单元格是与i连续相同的最后一个
                cell_next = self.sheet[letter + str(j)]
                if cell_next.value != cell.value:
                    j -= 1
                    break
                else:
                    j += 1
                if j > self.sheet.max_row:
                    j -= 1
                    break
            if j - i >= 1 and cell.value != '' and cell.value:
                # 如果有连续两格以上的单元格内容相同，进行融合
                msg = '%s%d:%s%d' % (letter, i, letter, j)
                self.sheet.merge_cells(msg)
            # 控制一下格式
            self.sheet[letter + str(i)].alignment = Alignment(horizontal='center',
                                                              vertical='top',
                                                              wrap_text=True)
            i = j + 1  # 继续指向下个单元格

    def get_product_all_list(self, f_name):
        # <0的波动率设置为绿色,f_name 是波动率列名，必填参数
        self.set_color(f_name, self.Color_green, self.rule_down)
        # >0的波动率设置为红色，f_name 是波动率列名，必填参数
        self.set_color(f_name, self.Color_red, self.rule_up)
        # 波动率显示为百分比，f_name 是波动率列名，必填参数
        self.set_percent(f_name)
        self.save()

    def rule_down(self, val):
        if val < 0:
            return True
        else:
            return False

    def rule_up(self, val):
        if val > 0:
            return True
        else:
            return False


def set_style_of_excel(product_list, file_name, product_name):
    df = pd.DataFrame(product_list)
    name_list = ['storeName', 'shardId', 'buyPrice', 'salePrice', 'transferTime', 'updateTime', 'fluctuate',
                 'transferCount', 'ownerNickName',
                 'fromUserName', 'activeCount', 'castQty']
    for item in df.columns:
        if item not in name_list:
            df.drop([item], axis=1)

    # df.drop(['hangingId', 'hangingNo','image','productId','saleStatus','creatorName','chainAccountAddress',
    # 'createTime','blindBoxLevelIcon',''], axis=1)
    df.rename(
        columns={'storeName': '藏品名称', 'shardId': '藏品编号', 'salePrice': '寄售价格', 'updateTime': '寄售时间', 'buyPrice': '购入价格',
                 'fluctuate': '涨幅', 'ownerNickName': '持有者昵称', 'fromUserName': '卖家昵称', 'transferTime': '购入时间',
                 'transferCount': '转手次数', 'activeCount': '流通量', 'castQty': '发行量'},
        inplace=True)
    # 定义列名顺序
    columns = ['藏品名称', '藏品编号', '购入价格', '寄售价格', '购入时间', '寄售时间', '涨幅', '转手次数', '持有者昵称',
               '卖家昵称', '流通量', '发行量']
    df = df[columns]
    # 初始化一个对象, 设定保存后的文件名和表名
    xlsx = XlsxSaver(columns, df, file_name, product_name + '-数据分析')
    # 传参：波动率的列名
    xlsx.get_product_all_list('涨幅')
